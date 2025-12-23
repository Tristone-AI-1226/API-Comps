import pandas as pd
import openpyxl
import io
import os
import requests
import json
import re
from urllib.parse import quote
from msal import ConfidentialClientApplication
import time
from google.api_core import exceptions
import google.generativeai as genai
from typing import List, Dict, Optional, Set

class GeminiCompanyExtractor:
    def __init__(self, source, api_key: str, target_company: str = None, max_sheets: int = 10, backup_api_key: str = None):
        """
        Initialize the extractor with Gemini integration.
        :param source: file path (str) or BytesIO stream (from SharePoint)
        :param api_key: Google Gemini API Key
        :param target_company: Name of the target company for context-aware extraction
        :param max_sheets: number of sheets to process
        :param backup_api_key: Optional backup API key for quota exhaustion
        """
        self.source = source
        self.target_company = target_company
        self.max_sheets = max_sheets
        self.api_key = api_key
        self.backup_api_key = backup_api_key
        self.results = {}
        
        # Configure Gemini
        if api_key:
             masked_key = api_key[:4] + "..." + api_key[-4:]
        else:
             print("[ERROR] Gemini API Key is MISSING!")
        genai.configure(api_key=api_key)

    def _convert_to_dataframe(self, workbook):
        """
        Convert Excel workbook to DataFrame, identifying both M&A and public comps sheets.
        Returns: (ma_sheet_data, public_sheet_data, has_ma, has_public)
        """
        ma_sheet_data = {}
        public_sheet_data = {}
        sheet_names = workbook.sheetnames
        
        # Patterns for different comp types
        # M&A pattern: Must have M&A/transaction/precedent/deal keywords
        # Matches: "M&A comps", "Transaction comps", "comps M&A", etc.
        # Does NOT match: just "comps" (that's public)
        ma_pattern = re.compile(r'(m&a|ma|transaction|precedent|deal|private).*comps|comps.*(m&a|ma|transaction|precedent|deal)', re.IGNORECASE)
        
        # Public pattern: equity/trading/public comps OR just "comps"
        # Matches: "Public comps", "Equity comps", "comps", etc.
        public_pattern = re.compile(r'(equity|trading|public).*comps|^comps$', re.IGNORECASE)
        
        ma_sheets = []
        public_sheets = []
        
        # Identify all matching sheets
        for name in sheet_names:
            # Skip hidden sheets
            if workbook[name].sheet_state != 'visible':
                continue

            name_stripped = name.strip()
            
            # Check M&A first (more specific)
            if ma_pattern.search(name_stripped):
                ma_sheets.append(name)
            # Then check public comps
            elif public_pattern.search(name_stripped):
                public_sheets.append(name)
        
        # Helper to score sheet names for prioritization
        def score_sheet_name(name):
            name_lower = name.lower()
            score = 100 # Base score
            
            # Penalize data source suffixes (with variations)
            data_sources = [
                'pitchbook', 'pitch book', 'pitch_book',
                'capiq', 'cap iq', 'cap_iq', 'capitaliq', 'capital iq', 'capital_iq',
                'factset', 'fact set', 'fact_set',
                'crunchbase', 'crunch base', 'crunch_base',
                'preqin', 'pre qin', 'pre_qin'
            ]
            
            for source in data_sources:
                if source in name_lower:
                    score -= 50
                    break  # Only penalize once
            
            # Boost sector/industry keywords
            sector_keywords = [
                'sector', 'industry', 'domain', 'segment', 'category'
            ]
            
            for keyword in sector_keywords:
                if keyword in name_lower:
                    score += 30
                    break  # Only boost once
            
            # Prefer shorter names (usually "Public Comps" vs "Public Comps_Pitchbook")
            score -= len(name)
            
            return score

        # Sort and limit to top 2 sheets
        if ma_sheets:
            ma_sheets_scored = [(s, score_sheet_name(s)) for s in ma_sheets]
            ma_sheets_scored.sort(key=lambda x: x[1], reverse=True)
            ma_sheets = [s for s, _ in ma_sheets_scored[:2]]
        
        if public_sheets:
            public_sheets_scored = [(s, score_sheet_name(s)) for s in public_sheets]
            public_sheets_scored.sort(key=lambda x: x[1], reverse=True)
            public_sheets = [s for s, _ in public_sheets_scored[:2]]
        
        # Process M&A sheets
        for sheet_name in ma_sheets:
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                row_list = [str(cell) if cell is not None else "" for cell in row]
                data.append(row_list)
            df = pd.DataFrame(data)
            ma_sheet_data[sheet_name] = df
        
        # Process public comps sheets
        for sheet_name in public_sheets:
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                row_list = [str(cell) if cell is not None else "" for cell in row]
                data.append(row_list)
            df = pd.DataFrame(data)
            public_sheet_data[sheet_name] = df
        
        # Fallback: if no specific sheets found, use first 3 visible sheets as public comps
        if not ma_sheets and not public_sheets:
            # No comps sheets found - use first 3 visible sheets as fallback
            visible_sheets = [n for n in sheet_names if workbook[n].sheet_state == 'visible']
            for sheet_name in visible_sheets[:3]:
                sheet = workbook[sheet_name]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    row_list = [str(cell) if cell is not None else "" for cell in row]
                    data.append(row_list)
                df = pd.DataFrame(data)
                public_sheet_data[sheet_name] = df
        
        has_ma = len(ma_sheet_data) > 0
        has_public = len(public_sheet_data) > 0
        
        return ma_sheet_data, public_sheet_data, has_ma, has_public

    def _prepare_context_for_gemini(self, sheet_data, max_chars=3200000):
        """
        Prepare Excel data as structured text context for Gemini.
        """
        context = "Below is data from an Excel file containing company information:\n\n"
        
        for sheet_name, df in sheet_data.items():
            sheet_context = f"=== SHEET: {sheet_name} ===\n"
            sheet_str = df.head(50).to_json(orient='records')
            
            if len(context) + len(sheet_context) + len(sheet_str) > max_chars:
                remaining_chars = max_chars - len(context) - len(sheet_context)
                if remaining_chars > 0:
                    sheet_str = sheet_str[:remaining_chars] + "\n... (truncated due to size limits)"
                else:
                    break  # Context limit reached 
            
            sheet_context += sheet_str
            sheet_context += "\n\n"
            context += sheet_context
        
        if len(context) > max_chars:
            context = context[:max_chars] + "\n... (truncated due to size limits)"
        
        return context

    def extract_with_gemini(self, model_name='gemini-2.5-flash'):
        """
        Extract data from Excel using Gemini in a single unified call.
        Includes robust error handling for 503 (Service Unavailable) and 429 (Resource Exhausted).
        """
        try:
            # Load workbook
            if isinstance(self.source, io.BytesIO):
                wb = openpyxl.load_workbook(self.source, data_only=True)
            else:
                file_ext = os.path.splitext(self.source)[1].lower()
                if file_ext == ".csv":
                    df = pd.read_csv(self.source)
                    buf = io.BytesIO()
                    df.to_excel(buf, index=False)
                    buf.seek(0)
                    wb = openpyxl.load_workbook(buf, data_only=True)
                else:
                    wb = openpyxl.load_workbook(self.source, data_only=True)
        except Exception as e:
            print(f"ERROR loading workbook: {e}")
            raise

        # Detect both M&A and public comps sheets
        ma_sheet_data, public_sheet_data, has_ma, has_public = self._convert_to_dataframe(wb)
        
        combined_sheet_data = {}
        combined_sheet_data.update(ma_sheet_data)
        combined_sheet_data.update(public_sheet_data)
        
        if not combined_sheet_data:
             wb.close()
             return None

        # Prepare context
        context = self._prepare_context_for_gemini(combined_sheet_data)
        wb.close()

        # Build Unified Prompt
        target_context = ""
        if self.target_company:
            target_context = f"TARGET COMPANY: {self.target_company}\n"

        prompt = f"""{context}

{target_context}TASK: Analyze the provided Excel data to extract competitive intelligence for {self.target_company if self.target_company else "the target company"}.

You must perform THREE tasks in a single pass:
1. Extract M&A Transactions
2. Extract Public Comparable Companies
3. Classify all entities

INSTRUCTIONS:

1. M&A TRANSACTIONS
   - Look for transaction/deal lists (Target, Acquirer, Deal Value, etc.).
   - Extract up to 10 most relevant transactions.
   - Ignore summary rows (Total, Average, Mean, Median).
   - Fields to Extract:
     * Target (Target, Company)
     * Acquirer (Acquirer, Buyer, Bidder)
     * Type (Deal Type, Description)
     * Metrics (Revenue, Valuation/EV, EV/Revenue, EV/EBITDA) - Keep units/currency.
   - CLASSIFY ACQUISITION TYPE:
     * "Strategic": Acquirer in SAME/Adjacent industry as Target.
     * "Financial": Acquirer is PE fund, investment firm, or financial sponsor.
     * "Unknown": Insufficient info.

2. PUBLIC COMPS
   - Look for lists of comparable public companies.
   - Extract public companies similar to {self.target_company if self.target_company else "the target"}.
   - CLASSIFY EACH COMPANY:
     * "Verified": Score >= 70 (Direct competitor, same market/products).
     * "To Cross-Check": Score < 70 (Indirect, substitute, or unclear).
   - LIMITS:
     * Max 10 Verified Competitors.
     * Max 10 To Cross-Check.

OUTPUT JSON FORMAT:
{{
  "ma_transactions": [
    {{
      "target": "Target Name",
      "acquirer": "Acquirer Name",
      "type": "Deal Type",
      "acquisition_type": "Strategic|Financial|Unknown",
      "revenue": "100M",
      "valuation": "500M",
      "ev_revenue": "5.0x",
      "ev_ebitda": "12.5x"
    }},
    ...
  ],
  "public_comps": {{
    "verified": [
      {{"name": "Company A", "score": 95, "reason": "Direct competitor in X space"}},
      ...
    ],
    "to_crosscheck": [
      {{"name": "Company B", "score": 40, "reason": "Different industry sector"}}
    ]
  }}
}}

CRITICAL: Provide ONLY valid JSON. No markdown formatting.
"""
        
        current_model_name = model_name
        fallback_model_name = 'gemini-2.5-flash-lite'
        
        # State Tracking
        retries = 0
        used_backup_key = False
        
        # Hard limits on attempts
        # We model this as a loop where we decide next action: Retry, Backup, Fail
        attempts = 0
        max_attempts = 4 # Enough for initial + retries
        
        while attempts < max_attempts:
            attempts += 1
            try:
                model = genai.GenerativeModel(current_model_name)
                response = model.generate_content(prompt)
                full_response = response.text
                break # Success!
                
            except exceptions.ServiceUnavailable:
                # 503 Strategy: Retry 1 (Same) -> Retry 2 (Fallback) -> Fail
                # We reuse the logic from previous task, but integrated here
                # Simplified:
                # If attempt 1 -> Wait 2s -> Retry Same
                # If attempt 2 -> Wait 2s -> Switch Model -> Retry Fallback
                # If attempt 3 -> Fail
                
                if attempts == 1:
                    time.sleep(2)
                    continue
                elif attempts == 2:
                    current_model_name = fallback_model_name
                    time.sleep(2)
                    continue
                else:
                    print("[ERROR] 503 Service Unavailable - All retries exhausted.")
                    return None
            
            except exceptions.ResourceExhausted as e:
                err_str = str(e)
                
                # Check for Quota Limit (PerDay)
                if "PerDay" in err_str or "Quota" in err_str:
                    print("[INFO] 429 Type: Quota Limit (PerDay)")
                    if self.backup_api_key and not used_backup_key:
                        print("[INFO] Switching to BACKUP API KEY.")
                        genai.configure(api_key=self.backup_api_key)
                        used_backup_key = True
                        # Retry immediately with backup key (same model)
                        continue
                    else:
                        print("[ERROR] Quota limit hit and no backup key available (or already used). Failing.")
                        return None
                        
                else: 
                    # Default to Rate Limit (PerMinute)
                    print("[INFO] 429 Type: Rate Limit (PerMinute)")
                    # Strategy:
                    # 1. Switch to Fallback Model (gemini-2.5-flash-lite) immediately
                    # 2. Retry
                    # 3. If fail -> Error
                    
                    if attempts == 1:
                        current_model_name = fallback_model_name
                        continue
                    else:
                         print("[ERROR] Rate limit retries exhausted (fallback model failed). Failing.")
                         return None
                         
            except Exception as e:
                print(f"ERROR in unified extraction: {e}")
                return None
        else:
            # Loop finished without break
            return None

        try:
            # clean response
            
            # clean response
            full_response = full_response.replace('```json', '').replace('```', '').strip()
            json_start = full_response.find('{')
            json_end = full_response.rfind('}') + 1
            
            if json_start == -1 or json_end == 0:
                print("ERROR: No JSON found in response")
                return None
                
            json_str = full_response[json_start:json_end]
            result = json.loads(json_str)
            
            # Post-process to ensure structure matches what main.py expects (partially)
            # We will return the raw unified result, and main.py will handle aggregation.
            
            # Add metadata
            result['type'] = 'unified' # Signal to caller
            
            # Capture usage metadata
            usage = response.usage_metadata
            usage_stats = {
                "prompt_token_count": usage.prompt_token_count,
                "candidates_token_count": usage.candidates_token_count,
                "total_token_count": usage.total_token_count,
                "input_char_count": len(prompt)
            }
            result['usage'] = usage_stats

            # Stats (optional debugging)
            ma_count = len(result.get('ma_transactions', []))
            pub_verified_count = len(result.get('public_comps', {}).get('verified', []))
            pub_check_count = len(result.get('public_comps', {}).get('to_crosscheck', []))
            
            return result

        except Exception as e:
            print(f"ERROR in unified extraction: {e}")
            print(f"Full response was: {full_response if 'full_response' in locals() else 'N/A'}")
            return None


class CopilotResponseProcessor:
    def __init__(self, copilot_response: str, target_company: str, api_key: str):
        """
        Initialize processor with copilot response and target company name.
        """
        self.copilot_response = copilot_response
        self.target_company = target_company
        self.api_key = api_key
        self.file_paths = []
        self.relative_paths = []
        self.all_competitors = set()
        self.verified_competitors = []
        self.to_crosscheck = []
        self.file_wise_companies = {}
        
        # Configure Gemini
        genai.configure(api_key=api_key)

    def extract_file_paths(self):
        """Extract unique file paths from copilot response and apply filtering logic."""
        # Pattern to match "Full Path: " followed by the path, cut at file extraction
        pattern = r'Full Path:\s*(.+?\.(?:xlsx|xls|csv|pptx|pdf))'
        matches = re.findall(pattern, self.copilot_response, re.IGNORECASE)
        unique_paths = list(set([path.strip() for path in matches]))
        
        # Apply filtering and balancing logic
        self.file_paths = self._filter_and_balance_files(unique_paths)

        # Extract relative paths starting from "Shared Documents/"
        self.relative_paths = []
        for path in self.file_paths:
            if "Shared Documents/" in path:
                relative_path = path.split("Shared Documents/", 1)[1]
                self.relative_paths.append(relative_path)
            else:
                self.relative_paths.append(path)

        return self.file_paths, self.relative_paths

    def _filter_and_balance_files(self, file_paths: List[str]) -> List[str]:
        """
        Filter files to max 4, balancing M&A and Public Comps.
        Priority: Folder path -> Filename Regex.
        
        Logic:
        - If both exist: Top 2 M&A + Top 2 Public
        - If only one exists: Top 4 of that type
        """
        ma_files = []
        public_files = []
        
        # Regex patterns
        ma_pattern = re.compile(r'(m&a|ma|transaction|precedent|deal|private).*comps|comps.*(m&a|ma|transaction|precedent|deal)', re.IGNORECASE)
        public_pattern = re.compile(r'(equity|trading|public).*comps|^comps', re.IGNORECASE)
        
        for path in file_paths:
            # 1. Folder Path Check (Priority)
            # Normalizing path separators just in case
            norm_path = path.replace('\\', '/')
            
            if "Relative Valuation/M&A" in norm_path or "Relative Valuation/MA" in norm_path:
                 ma_files.append(path)
                 continue
            elif "Relative Valuation/Public" in norm_path:
                 public_files.append(path)
                 continue
                 
            # 2. Filename Regex Check (Fallback)
            filename = os.path.basename(path)
            if ma_pattern.search(filename):
                ma_files.append(path)
            elif public_pattern.search(filename):
                public_files.append(path)
            else:
                # Default fallback if "comps" is in name
                if "comps" in filename.lower():
                     public_files.append(path)
        
        # Selection Logic
        final_files = []
        
        # Check if both categories exist
        if ma_files and public_files:
            # Take up to 2 from each
            final_files.extend(ma_files[:2])
            final_files.extend(public_files[:2])
            
        elif ma_files:
            # Only M&A
            final_files.extend(ma_files[:4])
            
        elif public_files:
            # Only Public
            final_files.extend(public_files[:4])
        
        if not final_files and file_paths:
             # If no files matched any pattern but paths exist, take top 4 raw to be safe
             final_files = file_paths[:4]
            
        return sorted(list(set(final_files)))

    def aggregate_unified_results(self, all_extraction_results: List[dict]):
        """
        Aggregate results from multiple files (unified extraction).
        - Consolidated M&A transactions.
        - De-duplicate and sort verified competitors.
        - De-duplicate and sort to-crosscheck competitors.
        """
        all_ma = []
        verified_map = {} # name -> {score, reason}
        crosscheck_map = {} # name -> {score, reason}
        
        for result in all_extraction_results:
            if not result:
                continue
                
            # Aggregate M&A
            if 'ma_transactions' in result:
                all_ma.extend(result['ma_transactions'])
                
            # Aggregate Public Comps
            if 'public_comps' in result:
                # Verified
                for comp in result['public_comps'].get('verified', []):
                    name = comp.get('name')
                    if name:
                        # Keep the one with higher score if duplicate
                        if name not in verified_map or comp.get('score', 0) > verified_map[name].get('score', 0):
                             verified_map[name] = comp
                             
                # To Cross-check
                for comp in result['public_comps'].get('to_crosscheck', []):
                    name = comp.get('name')
                    if name:
                        if name not in crosscheck_map or comp.get('score', 0) > crosscheck_map[name].get('score', 0):
                             crosscheck_map[name] = comp
        
        # Final Processing
        
        # 1. Verification vs Crosscheck Conflict Resolution
        # If a company is in both, prioritize Verified
        for name in list(crosscheck_map.keys()):
            if name in verified_map:
                del crosscheck_map[name]
                
        # 2. Sort Lists
        self.verified_competitors = sorted(verified_map.values(), key=lambda x: x.get('score', 0), reverse=True)
        self.to_crosscheck = sorted(crosscheck_map.values(), key=lambda x: x.get('score', 0), reverse=True)
        
        # 3. Apply global limits (optional, but individual file limits are already 10)
        # We'll keep all for now to maximize recall across files, or limit if total is huge.
        # Let's limit to top 20 total for now to keep output clean? 
        # Plan didn't specify global limit, only "Max 10 per category" which was per file prompt.
        # But user requested "Output Bounds... Gemini output is explicitly capped... Results are ranked...".
        # Let's stick to the aggregated list as is, maybe top 20 is safe.
        self.verified_competitors = self.verified_competitors[:20] 
        self.to_crosscheck = self.to_crosscheck[:20]
        
        return {
            "ma_transactions": all_ma,
            "verified_competitors": self.verified_competitors,
            "to_crosscheck": self.to_crosscheck,
            "verified_count": len(self.verified_competitors),
            "crosscheck_count": len(self.to_crosscheck),
            "ma_count": len(all_ma)
        }

def get_graph_token(tenant_id, client_id, client_secret):
    """Get Microsoft Graph API access token."""
    authority = f"https://login.microsoftonline.com/{tenant_id}"
    try:
        app = ConfidentialClientApplication(
            client_id, authority=authority, client_credential=client_secret)
        token = app.acquire_token_for_client(
            scopes=["https://graph.microsoft.com/.default"])
        if "access_token" in token:
            return token["access_token"]
        else:
            error_desc = token.get("error_description", "No error description")
            error_code = token.get("error", "Unknown error")
            raise Exception(f"No access token in response. Error: {error_code} - {error_desc}")
    except Exception as e:
        raise Exception(f"Error getting token: {e}")

def search_file_by_name(access_token, drive_id, filename):
    """Search for a file by name in SharePoint if direct path fails."""
    search_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/search(q='{filename}')"
    headers = {"Authorization": f"Bearer {access_token}"}
    try:
        resp = requests.get(search_url, headers=headers)
        resp.raise_for_status()
        results = resp.json()
        if "value" in results and len(results["value"]) > 0:
            return results["value"][0]
        return None
    except Exception:
        return None

def download_file_from_sharepoint(access_token, drive_id, relative_path):
    """Download file from SharePoint using Microsoft Graph API with fallback search."""
    # Try direct path first
    encoded_path = quote(relative_path, safe='/')
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded_path}:/content"
    headers = {"Authorization": f"Bearer {access_token}"}

    try:
        resp = requests.get(url, headers=headers)

        if resp.status_code == 404:
            filename = os.path.basename(relative_path)
            file_item = search_file_by_name(access_token, drive_id, filename)

            if file_item:
                download_url = file_item.get('@microsoft.graph.downloadUrl')
                if download_url:
                    resp = requests.get(download_url)
                    resp.raise_for_status()
                    return io.BytesIO(resp.content)
                else:
                    raise Exception(f"No download URL for file: {filename}")
            else:
                raise Exception(f"File not found: {filename}")

        resp.raise_for_status()
        return io.BytesIO(resp.content)

    except Exception as e:
        raise Exception(f"Download failed: {e}")
